"""
エクセルファイルを読み込み、棒グラフを作成するスクリプト

dl_censuspopdata.xlsxの内容

CensusTract	State	County	POP2010
05149952301	AR	Yell	3126
05149952302	AR	Yell	2941
05149952401	AR	Yell	3180
05149952402	AR	Yell	5515
05149952500	AR	Yell	4796
05149952600	AR	Yell	2627
06001400100	CA	Alameda	2937
06001400200	CA	Alameda	1974
06001400300	CA	Alameda	4865
06001400400	CA	Alameda	3703
06001400500	CA	Alameda	3517
06001400600	CA	Alameda	1571
06001400700	CA	Alameda	4206
"""

"""
上記の形式のExcelファイル（dl_censuspopdata.xlsx）を読み込み
各州（State）の人口（POP2010）を集計し、
別のシートを作成し、棒グラフで表示するスクリプトを作成する。

ch13/readCensusExcel.pyを参考にして作成しました。
"""

from datetime import datetime

import openpyxl
from openpyxl.chart import BarChart, Reference


def run():
    # Excelファイルを読み込む
    print("ワークブックを開いています...")
    wb = openpyxl.load_workbook("dl_censuspopdata.xlsx")
    sheet = wb["Population by Census Tract"]

    # 各州の人口を集計
    print("各州の人口を集計しています...")
    population_data = {}

    # 2行目から最終行までを読み込む
    for row in range(2, sheet.max_row + 1):
        # B列（州名）とD列（人口）を読み込む
        state = sheet[f"B{row}"].value
        pop = sheet[f"D{row}"].value
        population_data.setdefault(state, 0)
        population_data[state] += pop

    # population_dataを人口の降順でソート
    population_data = dict(
        sorted(population_data.items(), key=lambda x: x[1], reverse=True)
    )

    # シートを作成 13.5.2
    print("新しいシートを作成しています...")
    new_sheet = wb.create_sheet(title="Population")
    new_sheet["A1"] = "State"
    new_sheet["B1"] = "Population"

    # シートを切り替える
    wb.active = new_sheet

    # 集計結果を新しいシートに書き込む
    for i, (state, pop) in enumerate(population_data.items(), start=2):
        new_sheet[f"A{i}"] = state
        new_sheet[f"B{i}"] = pop

    # 棒グラフを作成
    print("棒グラフを作成しています...")
    print(len(population_data))

    # 参照するデータの範囲を指定, p.398を参照。
    # 以下のコードを理解してみましょう。
    # 余裕があれば、別のグラフを作成してみましょう。
    # https://openpyxl.readthedocs.io/en/stable/charts/bar.html
    states = Reference(new_sheet, min_col=1, min_row=2, max_row=len(population_data))
    populations = Reference(
        new_sheet, min_col=2, min_row=1, max_col=2, max_row=len(population_data)
    )
    # 棒グラフにデータを設定
    chart = BarChart()
    chart.add_data(populations, titles_from_data=True)
    chart.set_categories(states)

    # グラフのタイトルと軸ラベルを設定
    chart.title = "Population by State"
    chart.x_axis.title = "State"
    chart.y_axis.title = "Population"

    # グラフをシートに追加
    new_sheet.add_chart(chart, "E1")

    # Excelファイルを保存
    file_name = f'population-{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(file_name)

    print(f"{file_name}を保存しました。")


if __name__ == "__main__":
    run()
